#!/usr/bin/env python3
"""YAML → Excel register table (same columns/layout as xml_to_excel)."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

from yml_model import default_output_dir, load_yml_model, stem_name


def _write_reg_sheet(ws, all_items: list) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    reg_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

    headers = [
        "register_name",
        "address",
        "field_name",
        "bit offset",
        "access",
        "default",
        "description",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row = 2
    for item in all_items:
        num_fields = len(item["fields"])
        start_row = current_row

        ws.cell(row=start_row, column=1, value=item["name"])
        ws.cell(row=start_row, column=2, value=item["offset"])
        ws.cell(row=start_row, column=5, value=item["access"])
        ws.cell(row=start_row, column=6, value=item["reset"])
        ws.cell(row=start_row, column=7, value=item["description"])

        for col in range(1, 8):
            cell = ws.cell(row=start_row, column=col)
            cell.border = thin_border
            cell.fill = reg_fill
            cell.alignment = center_align if col in (1, 2, 5, 6) else left_align

        if num_fields == 0:
            current_row += 1
            continue

        end_row = start_row + num_fields
        for i, f in enumerate(item["fields"]):
            row = start_row + 1 + i
            end_bit = f["bit_offset"] + f["bit_width"] - 1
            if f["bit_width"] == 1:
                bit_range = f"[{f['bit_offset']}:{f['bit_offset']}]"
            else:
                bit_range = f"[{end_bit}:{f['bit_offset']}]"
            ws.cell(row=row, column=3, value=f["name"])
            ws.cell(row=row, column=4, value=bit_range)
            ws.cell(row=row, column=5, value=f["access"])
            ws.cell(row=row, column=6, value=item["reset"])
            ws.cell(row=row, column=7, value=f["description"])
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align if col in (1, 2, 4, 5, 6) else left_align

        current_row = end_row + 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 45
    ws.freeze_panes = "A2"


def generate_excel(model: dict, out_path: Path) -> Path:
    wb = openpyxl.Workbook()
    # Info sheet (component mode compatible)
    info = wb.active
    info.title = f"{stem_name(model).lower()}_sc"[:31]
    info.append(
        [
            "name",
            "version",
            "block_name",
            "base_address",
            "range",
            "protocol",
            "description",
        ]
    )
    info.append(
        [
            model["component_name"],
            model["version"],
            model["block_name"],
            model["base_address"],
            model["range"],
            model["protocol"],
            model["description"],
        ]
    )

    # Register sheet
    reg_ws = wb.create_sheet(f"{stem_name(model).lower()}_sc_reg"[:31])
    items = list(model["registers"])
    # Flatten interrupt expanded registers for table view (like xml_to_excel)
    for intp in model["interrupts"]:
        items.extend(intp["registers"])
    items.sort(key=lambda r: int(r["offset"], 0))
    _write_reg_sheet(reg_ws, items)

    if model["interrupts"]:
        intp_ws = wb.create_sheet(f"{stem_name(model).lower()}_sc_intp"[:31])
        intp_items = []
        for intp in model["interrupts"]:
            intp_items.append(
                {
                    "name": intp["name"],
                    "offset": intp["base_offset"],
                    "access": intp["access"],
                    "reset": intp["reset"],
                    "description": intp["description"],
                    "fields": intp["fields"],
                }
            )
        _write_reg_sheet(intp_ws, intp_items)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Convert yml2reg YAML to Excel register table")
    parser.add_argument("yaml_file", help="Input YAML register description")
    parser.add_argument("-o", "--output-dir", default="", help="Output directory (default: beside YAML)")
    args = parser.parse_args(argv)

    yaml_path = Path(args.yaml_file)
    model = load_yml_model(yaml_path)
    out_dir = default_output_dir(yaml_path, args.output_dir or None)
    out_path = out_dir / f"{stem_name(model)}.xlsx"
    generate_excel(model, out_path)
    print(f"Generated: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
