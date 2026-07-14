#!/usr/bin/env python3
"""Rebuild the reviewed two-domain workbook from the upf-gen template."""

from pathlib import Path
import os
import shutil
import openpyxl


HERE = Path(__file__).resolve().parent
PROJECT_ROOT = Path(os.environ.get("PROJECT_ROOT", HERE.parents[4]))
TEMPLATE = Path(os.environ.get(
    "UPF_GEN_TEMPLATE",
    PROJECT_ROOT / ".agents" / "skills" / "upf-gen" / "assets" / "power_intent_filled.xlsx",
))
OUTPUT = HERE / "power_intent.xlsx"

shutil.copyfile(TEMPLATE, OUTPUT)
book = openpyxl.load_workbook(OUTPUT)


def replace_rows(sheet_name, rows):
    sheet = book[sheet_name]
    for row in sheet.iter_rows(min_row=4):
        for cell in row:
            cell.value = None
    for row_index, values in enumerate(rows, start=4):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)


replace_rows("Supplies", [
    ("VDD_AO", "primary", None, 1.8, None, None, "Y", "PD_AO primary"),
    ("VDD_PLL", "primary", None, 1.8, None, None, "Y", "PD_AO extra_supplies_1"),
    ("VDD_MEM", "primary", None, 1.8, None, None, "Y", "PD_AO extra_supplies_2"),
    ("VDDIO", "primary", None, 3.3, None, None, "Y", "PD_AO extra_supplies_3"),
    ("VDD_SW_IN", "primary", None, 1.2, None, None, "Y", "PD_SW switch input"),
    ("VDD_SW", "switched", "VDD_SW_IN", 1.2, None, None, "N", "PD_SW switched output"),
])

replace_rows("Domains", [
    ("PD_AO", ".", "VDD_AO", "N", "N", None, None,
     "AO extent: u_aon_ctrl, u_pll_macro, u_sram_macro, u_pad_in, u_pad_out; PLL/MEM/IO use extra supplies"),
    ("PD_SW", "u_sw_core", "VDD_SW", "Y", "N", None, "0",
     "Only switchable digital core"),
])

replace_rows("PowerStates", [
    ("ALL_ON", "ON", "ON", "All domains powered"),
    ("SW_OFF", "ON", "OFF", "PD_SW off; PD_AO remains on"),
])
power_states = book["PowerStates"]
power_states.cell(row=3, column=2, value="PD_AO")
power_states.cell(row=3, column=3, value="PD_SW")
for column in range(4, 10):
    power_states.cell(row=3, column=column, value=None)

replace_rows("Isolation_LS", [
    ("isolation", "PD_SW", "PD_AO", "0", "self", "sw_iso_n",
     "Clamp switchable responses at the PD_SW boundary; active-low release"),
    ("level_shifter", "PD_SW", "PD_AO", "auto", "self", None,
     "1.2 V to 1.8 V responses; co-located with isolation"),
])

control = book["Control"]
control["B4"] = "u_aon_ctrl"
control["B5"] = "sw_iso_n"
control["B6"] = "active_low"
control["B7"] = "sw_en"
control["B8"] = "unused_ret_save"
control["B9"] = "unused_ret_restore"

book.save(OUTPUT)
print(OUTPUT)
